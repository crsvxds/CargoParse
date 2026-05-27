import os
import sys
import random
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment


class CustomsApp:

    def __init__(self, root):
        self.root = root
        self.root.title("Customs Data Consolidator (ВЭД Китай)")
        self.root.geometry("750x640")
        self.root.minsize(700, 570)

        # Список шуточных фраз для статуса загрузки
        self.funny_statuses = [
            "Разгружаем контейнер из Гуанчжоу...",
            "Проверяем таможенную декларацию...",
            "Подкупаем инспектора шоколадкой...",
            "Завариваем крепкий кофе...",
            "Пересчитываем коробки вручную...",
            "Ищем потерявшийся артикул под столом...",
            "Ждем, пока китайская сторона подпишет доки...",
            "Сортируем маркировку левой пяткой...",
            "Проходим досмотр без регистрации и СМС...",
            "Запускаем дроны над складом..."
        ]

        # Переменные путей (вернули дефолтные значения, как было)
        self.articles_path = tk.StringVar(value="articles.txt")
        self.codes_dir = tk.StringVar(value="codes")
        self.output_path = tk.StringVar(value="result.xlsx")
        
        # Чекбокс для автооткрытия файла
        self.auto_open = tk.BooleanVar(value=True)

        self.create_widgets()

    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- БЛОК ВЫБОРА ФАЙЛОВ ---
        file_frame = ttk.LabelFrame(main_frame, text=" Настройки путей ", padding="10")
        file_frame.pack(fill=tk.X, pady=5)

        ttk.Label(file_frame, text="Файл артикулов:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(file_frame, textvariable=self.articles_path, width=55).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(file_frame, text="Обзор...", command=self.browse_articles).grid(row=0, column=2, pady=5)

        ttk.Label(file_frame, text="Папка с кодами:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(file_frame, textvariable=self.codes_dir, width=55).grid(row=1, column=1, padx=5, pady=5)
        ttk.Button(file_frame, text="Обзор...", command=self.browse_codes).grid(row=1, column=2, pady=5)

        ttk.Label(file_frame, text="Сохранить результат как:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(file_frame, textvariable=self.output_path, width=55).grid(row=2, column=1, padx=5, pady=5)
        ttk.Button(file_frame, text="Обзор...", command=self.browse_output).grid(row=2, column=2, pady=5)

        # --- БЛОК ЛОГОВ И ПРОГРЕССА ---
        log_frame = ttk.LabelFrame(main_frame, text=" Лог выполнения ", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        self.log_text = tk.Text(log_frame, height=12, width=85, state=tk.DISABLED)
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)

        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Индикатор прогресса и статусная строка
        progress_frame = ttk.Frame(main_frame, padding="5")
        progress_frame.pack(fill=tk.X, pady=5)
        
        self.progress_bar = ttk.Progressbar(progress_frame, orient=tk.HORIZONTAL, mode='determinate')
        self.progress_bar.pack(fill=tk.X, pady=(0, 5))
        
        status_sub_frame = ttk.Frame(progress_frame)
        status_sub_frame.pack(fill=tk.X)
        
        self.status_label = ttk.Label(status_sub_frame, text="Готов к работе", font=("Segoe UI", 9, "italic"))
        self.status_label.pack(side=tk.LEFT)
        
        self.progress_label = ttk.Label(status_sub_frame, text="0%", width=6, anchor=tk.E)
        self.progress_label.pack(side=tk.RIGHT)

        # --- КНОПКИ УПРАВЛЕНИЯ И ЧЕКБОКСЫ ---
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=10)

        self.run_btn = ttk.Button(btn_frame, text="СТАРТ ОБРАБОТКИ", command=self.start_processing_thread)
        self.run_btn.pack(side=tk.LEFT, padx=5, ipadx=10, ipady=5)

        ttk.Checkbutton(btn_frame, text="Открыть Excel после завершения", variable=self.auto_open).pack(side=tk.LEFT, padx=15)

        self.save_log_btn = ttk.Button(btn_frame, text="Сохранить лог в .txt", command=self.save_log_to_file)
        self.save_log_btn.pack(side=tk.RIGHT, padx=5, ipadx=5, ipady=5)

    def browse_articles(self):
        file = filedialog.askopenfilename(filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")])
        if file: self.articles_path.set(file)

    def browse_codes(self):
        directory = filedialog.askdirectory()
        if directory: self.codes_dir.set(directory)

    def browse_output(self):
        file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if file: self.output_path.set(file)

    def update_status_phrases_loop(self):
        if getattr(self, 'is_running', False):
            random_phrase = random.choice(self.funny_statuses)
            self.status_label.configure(text=random_phrase)
            self.root.after(2000, self.update_status_phrases_loop)

    def update_progress(self, current, total):
        if total == 0:
            percent = 0
        else:
            percent = int((current / total) * 100)
        self.progress_bar['value'] = percent
        self.progress_label.configure(text=f"{percent}%")
        self.root.update_idletasks()

    def log(self, message):
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)
        self.root.update_idletasks()

    def clear_log(self):
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.delete("1.0", tk.END)
        self.log_text.configure(state=tk.DISABLED)

    def save_log_to_file(self):
        log_content = self.log_text.get("1.0", tk.END).strip()
        if not log_content:
            messagebox.showwarning("Внимание", "Лог пуст. Нечего сохранять.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text Files", "*.txt")],
            initialfile="customs_verification_log.txt"
        )
        if file_path:
            try:
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write(log_content)
                messagebox.showinfo("Успех", f"Лог успешно сохранен в:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось сохранить файл:\n{e}")

    def start_processing_thread(self):
        self.is_running = True
        self.run_btn.configure(state=tk.DISABLED)
        
        self.update_status_phrases_loop()
        
        processing_thread = threading.Thread(target=self.start_processing)
        processing_thread.daemon = True
        processing_thread.start()

    def start_processing(self):
        self.clear_log()
        self.update_progress(0, 100)

        art_file = self.articles_path.get().strip()
        c_folder = self.codes_dir.get().strip()
        out_file = self.output_path.get().strip()

        errors = []
        if not art_file: errors.append("Не указан путь к файлу артикулов.")
        elif not os.path.exists(art_file): errors.append(f"Не найден файл артикулов: '{art_file}'")
        
        if not c_folder: errors.append("Не указан путь к папке с кодами.")
        elif not os.path.exists(c_folder): errors.append(f"Не найдена папка с кодами: '{c_folder}'")
        
        if not out_file: errors.append("Не указан путь для сохранения результата.")

        if errors:
            error_msg = "\n".join(errors)
            messagebox.showerror("Критическая ошибка", error_msg)
            self.log("[ОШИБКА ЗАПУСКА] Проверьте правильность заполнения путей.")
            self.is_running = False
            self.status_label.configure(text="Ошибка запуска")
            self.run_btn.configure(state=tk.NORMAL)
            return

        self.log("=== ЗАПУСК ВЕРИФИКАЦИИ ДЕКЛАРАЦИЙ ===")

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Результат"
            current_row = 1

            not_found_articles = []
            mismatched_articles = []
            duplicate_summary = {} 

            with open(art_file, "r", encoding="utf-8") as f:
                lines = [line.strip() for line in f if line.strip()]

            total_lines = len(lines)
            
            for index, line in enumerate(lines):
                parts = line.split()
                if len(parts) < 6:
                    self.log(f"[ПРОПУСК] Неверный формат строки: {line[:20]}...")
                    self.update_progress(index + 1, total_lines)
                    continue

                article = parts[0]
                unit = parts[-1]
                quantity = parts[-2]
                cargo_places = parts[-3]
                description = parts[-4]
                name = " ".join(parts[1:-4])

                try:
                    quantity_int = int(quantity)
                except ValueError:
                    self.log(f"[ОШИБКА] Неверное число количества у {article}")
                    self.update_progress(index + 1, total_lines)
                    continue

                possible_names = [article] + [f"{article}-{i}" for i in range(1, 11)]
                matched_files = []
                all_files = os.listdir(c_folder)

                for target in possible_names:
                    for file_name in all_files:
                        name_without_ext = os.path.splitext(file_name)[0]
                        if name_without_ext == target:
                            matched_files.append(os.path.join(c_folder, file_name))
                            break

                header_row = current_row
                ws[f"C{header_row}"] = article
                ws[f"D{header_row}"] = name
                ws[f"E{header_row}"] = description
                ws[f"F{header_row}"] = cargo_places
                ws[f"G{header_row}"] = quantity
                ws[f"H{header_row}"] = unit
                current_row += 1

                if not matched_files:
                    ws[f"A{current_row}"] = "НЕТУ"
                    not_found_articles.append(article)
                    self.log(f"Артикул {article}: Файлы декларации НЕ найдены")
                    current_row += 2
                    self.update_progress(index + 1, total_lines)
                    continue

                all_product_codes = []
                files_loaded_names = []
                dup_count_for_article = 0

                for file_path in matched_files:
                    try:
                        codes_df = pd.read_excel(file_path, header=None)
                        file_name_short = os.path.basename(file_path)
                        files_loaded_names.append(file_name_short)
                        
                        code_source_map = {}  
                        
                        for value in codes_df[0].tolist():
                            if pd.isna(value):
                                continue
                            cleaned_code = str(value).strip()
                            if cleaned_code:
                                if cleaned_code in code_source_map:
                                    dup_count_for_article += 1
                                    self.log(
                                        f"   [ДУБЛЬ ИГНОРИРОВАН] Артикул {article}:\n"
                                        f"     • Код: {cleaned_code}\n"
                                        f"     • Файл с повторением: {file_name_short}"
                                    )
                                else:
                                    code_source_map[cleaned_code] = file_name_short
                                    all_product_codes.append(cleaned_code)
                                    
                    except Exception as e:
                        self.log(f"[ОШИБКА ЧТЕНИЯ] файла {os.path.basename(file_path)}")

                if dup_count_for_article > 0:
                    duplicate_summary[article] = dup_count_for_article

                total_codes_count = len(all_product_codes)
                files_list_str = ", ".join(files_loaded_names)

                if total_codes_count != quantity_int:
                    diff = abs(quantity_int - total_codes_count)
                    status_text = "меньше" if total_codes_count < quantity_int else "больше"
                    ws[f"F{header_row + 1}"] = f"В файлах суммарно (без дублей): {total_codes_count}, {status_text.capitalize()} на {diff}"
                    
                    mismatched_articles.append({
                        "article": article, "expected": quantity_int, "actual": total_codes_count,
                        "diff": diff, "status": status_text, "files": files_list_str
                    })
                    self.log(f"Артикул {article}: РАСХОЖДЕНИЕ (Ожидалось {quantity_int}, собрано без дублей {total_codes_count})")
                else:
                    self.log(f"Артикул {article}: OK (Совпало {total_codes_count} шт.)")

                if total_codes_count == 0:
                    ws[f"A{current_row}"] = "НЕТУ"
                    current_row += 2
                    self.update_progress(index + 1, total_lines)
                    continue

                for code in all_product_codes:
                    ws[f"A{current_row}"] = code
                    current_row += 1
                current_row += 1
                
                self.update_progress(index + 1, total_lines)

            for col in ws.columns:
                max_length = max(len(str(cell.value or '')) for cell in col)
                column = col[0].column_letter
                for cell in col:
                    cell.alignment = Alignment(vertical="top")
                ws.column_dimensions[column].width = min(max_length + 5, 60)

            wb.save(out_file)

            # --- ФИНАЛЬНЫЙ ОТЧЕТ ---
            self.log("\n============================================================")
            self.log("ФИНАЛЬНЫЙ ОТЧЕТ ПО ОШИБКАМ И РАСХОЖДЕНИЯМ:")
            self.log("============================================================\n")

            if not_found_articles:
                self.log(f"[!] НЕ НАЙДЕНЫ ФАЙЛЫ ДЛЯ АРТИКУЛОВ ({len(not_found_articles)} шт.):")
                for art in not_found_articles:
                    self.log(f"    - Артикул {art}: файлы отсутствуют в папке.")
                self.log("")
            else:
                self.log("[✓] Для каждого артикула из списка найден хотя бы один файл.\n")

            if duplicate_summary:
                self.log(f"[!] ОБНАРУЖЕНЫ СОВПАДЕНИЯ (ДУБЛИКАТЫ) КОДОВ МАРКИРОВКИ:")
                for art, count in duplicate_summary.items():
                    self.log(f"    - Артикул {art}: отфильтровано {count} шт. повторяющихся кодов внутри файлов.")
                self.log("    *Повторяющиеся коды были полностью исключены из итогового Excel.*\n")
            else:
                self.log("[✓] Дубликатов кодов внутри файлов не обнаружено.\n")

            if mismatched_articles:
                self.log(f"[!] РАСХОЖДЕНИЕ ОБЩЕЙ СУММЫ КОДОВ ВНУТРИ ФАЙЛОВ ({len(mismatched_articles)} шт.):")
                for item in mismatched_articles:
                    self.log(
                        f"    - Артикул {item['article']}: "
                        f"В текстовике указано {item['expected']}, "
                        f"суммарно в файлах [{item['files']}] найдено (чистых) {item['actual']}. "
                        f"Итог: {item['status']} на {item['diff']} шт."
                    )
            else:
                self.log("[✓] Расхождений по количеству кодов во внутрянке не обнаружено.")
            
            self.log("\n============================================================")
            self.status_label.configure(text="Готово!")
            messagebox.showinfo("Успех", f"Обработка завершена!\nФайл сохранен: {out_file}")

            if self.auto_open.get():
                try:
                    os.startfile(out_file)
                except AttributeError:
                    import subprocess
                    subprocess.call(['open', out_file] if sys.platform == 'darwin' else ['xdg-open', out_file])

        except Exception as ex:
            messagebox.showerror("Критический сбой", f"Произошла ошибка во время сборки:\n{ex}")
            self.log(f"[КРАШ СИСТЕМЫ] {ex}")
            self.status_label.configure(text="Критический сбой")
        finally:
            self.is_running = False
            self.run_btn.configure(state=tk.NORMAL)


if __name__ == "__main__":
    root = tk.Tk()
    app = CustomsApp(root)
    root.mainloop()