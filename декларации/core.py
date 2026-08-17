import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment

# --- ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ ДЛЯ УМНОГО ПРЕОБРАЗОВАНИЯ В СТРОКУ (УБИРАЕТ .0) ---
def safe_str(val):
    if pd.isna(val):
        return ""
    # Если Pandas прочитал целое число как float (1200.0), делаем его целым (1200)
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()

# --- ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ ДЛЯ ОЧИСТКИ МУСОРА ---
def clean_raw_value(val):
    v = safe_str(val)
    if not v or v.lower() == 'nan' or v == 'НЕТУ': 
        return ""
    
    if 'код упаковки:' in v.lower():
        parts = v.split(':', 1)
        if len(parts) > 1:
            v = parts[1].strip()
            
    lower_v = v.lower()
    if 'код маркировки' in lower_v or 'код упаковки' in lower_v or 'артикул' in lower_v:
        return ""
        
    if not any(char.isdigit() for char in v):
        return ""
        
    return v

# --- ТОЧЕЧНОЕ УДАЛЕНИЕ СКОБОК ТОЛЬКО ДЛЯ 00, 01, 21 ---
def remove_specific_brackets(val):
    if pd.isna(val):
        return ""
    v_str = safe_str(val)
    if v_str.lower() == "nan" or v_str == "":
        return ""
    v_str = re.sub(r'\((00|01|21)\)', r'\1', v_str)
    return v_str

# =================================================================
# РЕЖИМ 1: ОСНОВНАЯ ВЕРИФИКАЦИЯ (Сверка по articles.txt)
# =================================================================
def process_customs_data(art_file, c_folder, out_dir, log_callback, progress_callback, stats_callback, check_boxes=True, remove_brackets=False):
    wb_import = Workbook()
    ws_import = wb_import.active
    ws_import.title = "УралИмпорт"
    curr_import = 1

    wb_trade = Workbook()
    ws_trade = wb_trade.active
    ws_trade.title = "УралТрейд"
    curr_trade = 1

    # Инициализация файла-рассписки
    wb_receipt = Workbook()
    ws_receipt = wb_receipt.active
    ws_receipt.title = "Рассписка"
    ws_receipt.append(["Артикул", "План (шт)", "Факт (шт)", "Принадлежность"])

    not_found_articles = []
    mismatched_articles = []
    duplicate_summary = {} 
    article_routing_summary = []
    global_seen_codes = {}

    with open(art_file, "r", encoding="utf-8") as f:
        lines = [line.strip() for line in f if line.strip()]

    # 1. АГРЕГАЦИЯ ТЕКСТОВОГО ФАЙЛА (Суммируем одинаковые артикулы)
    articles_dict = {}
    ok_counter = 0
    bad_counter = 0
    
    for line in lines:
        parts = line.split()
        if len(parts) < 6:
            art_name = parts[0] if parts else "НЕИЗВЕСТНО"
            log_callback(f"[ПРОПУСК] Неверный формат строки: {line[:20]}...")
            bad_counter += 1
            continue

        article = parts[0]
        unit = parts[-1]
        try:
            quantity_int = int(parts[-2])
        except ValueError:
            log_callback(f"[ОШИБКА] Неверное число количества у {article}")
            bad_counter += 1
            continue

        try:
            cargo_places_int = int(parts[-3])
        except ValueError:
            cargo_places_int = 0

        description = parts[-4]
        name = " ".join(parts[1:-4])

        if article not in articles_dict:
            articles_dict[article] = {
                'name': name,
                'desc': description,
                'qty': 0,
                'boxes': 0,
                'unit': unit
            }
        
        # Суммируем количество и коробки для дублирующихся строк
        articles_dict[article]['qty'] += quantity_int
        articles_dict[article]['boxes'] += cargo_places_int

    total_articles = len(articles_dict)
    stats_callback("total", total_articles)
    total_dups_counter = 0
    
    all_files = os.listdir(c_folder)
    unused_files = {f for f in all_files if f.endswith(('.xlsx', '.xls')) and not f.startswith('~')}

    # 2. УМНЫЙ ПОИСК ФАЙЛОВ ДЛЯ КАЖДОГО АРТИКУЛА
    file_to_article = {}
    for file_name in unused_files.copy():
        name_without_ext = os.path.splitext(file_name)[0]
        
        best_match = None
        for art in articles_dict.keys():
            if name_without_ext.startswith(art):
                if best_match is None or len(art) > len(best_match):
                    best_match = art
                    
        if best_match:
            if best_match not in file_to_article:
                file_to_article[best_match] = []
            file_to_article[best_match].append(file_name)
            unused_files.remove(file_name)

    # 3. ОСНОВНОЙ ЦИКЛ ОБРАБОТКИ
    for index, (article, adata) in enumerate(articles_dict.items()):
        name = adata['name']
        description = adata['desc']
        quantity_int = adata['qty']
        cargo_places_int = adata['boxes']
        unit = adata['unit']

        matched_files_names = file_to_article.get(article, [])
        matched_files_names.sort(key=lambda x: (os.path.splitext(x)[0] != article, x))
        matched_files = [os.path.join(c_folder, fn) for fn in matched_files_names]

        if not matched_files:
            not_found_articles.append(article)
            article_routing_summary.append((article, "❌ Файлы не найдены"))
            log_callback(f"[ПРОПУСК] ... Артикул {article}: Файлы декларации НЕ найдены (План: {quantity_int} шт)")
            
            # Запись в рассписку (когда нет файлов)
            ws_receipt.append([article, quantity_int, 0, "НЕТ ФАЙЛОВ"])
            
            bad_counter += 1
            stats_callback("bad", bad_counter)
            progress_callback(index + 1, total_articles)
            continue

        file_data_list = []
        for file_path in matched_files:
            try:
                df = pd.read_excel(file_path, header=None)
                # Извлекаем чистое имя без расширения .xlsx
                file_name_short = os.path.splitext(os.path.basename(file_path))[0]
                num_cols = df.shape[1]
                
                f_type = "import"
                f_codes_raw = []
                
                if num_cols >= 4:
                    f_type = "trade"
                    last_pack_code = "БЕЗ_КОДА_УПАКОВКИ"
                    for idx_row in range(len(df)):
                        mark_val = clean_raw_value(df.iloc[idx_row, 1])
                        raw_pack = df.iloc[idx_row, 3]
                        pack_val = safe_str(raw_pack)
                        
                        if pack_val and pack_val.lower() != 'nan' and not pack_val.lower().startswith('код'):
                            last_pack_code = pack_val
                            
                        if not mark_val or len(mark_val) < 10: continue
                        if remove_brackets: mark_val = remove_specific_brackets(mark_val)
                        if mark_val.startswith("(00)") or mark_val.startswith("00"): continue
                            
                        f_codes_raw.append((mark_val, last_pack_code))
                else:
                    if len(df) > 0:
                        for col_idx in range(df.shape[1]):
                            col_has_codes = False
                            for val in df.iloc[:, col_idx].dropna():
                                v = clean_raw_value(val)
                                if v and len(v) >= 10:
                                    col_has_codes = True
                                    if remove_brackets: v = remove_specific_brackets(v)
                                    f_codes_raw.append(v)
                            if col_has_codes: break
                                
                file_data_list.append({'name': file_name_short, 'type': f_type, 'raw_codes': f_codes_raw})
            except Exception as e:
                log_callback(f"[ОШИБКА ЧТЕНИЯ] файла {os.path.basename(file_path)}. Детали: {e}")

        trade_groups = {}  
        import_groups = {} 
        files_loaded_names = []
        dup_count_for_article = 0
        local_art_seen = set()
        
        trade_codes_count = 0
        import_codes_count = 0
        total_codes_count = 0
        actual_packages_set = set()

        for fdata in file_data_list:
            current_file_name = fdata['name']
            files_loaded_names.append(current_file_name)
            
            if fdata['type'] == 'trade':
                if current_file_name not in trade_groups: trade_groups[current_file_name] = {}
                for mark_val, pkg_code in fdata['raw_codes']:
                    if mark_val in local_art_seen or mark_val in global_seen_codes:
                        dup_count_for_article += 1
                        total_dups_counter += 1
                        stats_callback("dups", total_dups_counter)
                    else:
                        local_art_seen.add(mark_val)
                        global_seen_codes[mark_val] = current_file_name
                        trade_codes_count += 1
                        total_codes_count += 1
                        
                        if not pkg_code.startswith("БЕЗ_КОДА"): actual_packages_set.add(pkg_code)
                        if pkg_code not in trade_groups[current_file_name]: trade_groups[current_file_name][pkg_code] = []
                        trade_groups[current_file_name][pkg_code].append(mark_val)
            else:
                if current_file_name not in import_groups: import_groups[current_file_name] = []
                for mark_val in fdata['raw_codes']:
                    if mark_val in local_art_seen or mark_val in global_seen_codes:
                        dup_count_for_article += 1
                        total_dups_counter += 1
                        stats_callback("dups", total_dups_counter)
                    else:
                        local_art_seen.add(mark_val)
                        global_seen_codes[mark_val] = current_file_name
                        import_codes_count += 1
                        total_codes_count += 1
                        import_groups[current_file_name].append(mark_val)

        if dup_count_for_article > 0: duplicate_summary[article] = dup_count_for_article

        actual_packages_count = len(actual_packages_set)

        status_parts = []
        if trade_groups and import_groups:
            type_label = "Смешанный (Трейд+Импорт)"
            status_parts.append(f"Штук: план {quantity_int}, факт {total_codes_count} (Трейд: {trade_codes_count}, Импорт: {import_codes_count})")
            receipt_firm = f"СМЕШАННЫЙ (Трейд: {trade_codes_count}, Импорт: {import_codes_count})"
            if check_boxes and cargo_places_int > 0:
                status_parts.append(f"Коробок(Трейд): план {cargo_places_int}, факт {actual_packages_count}")
        elif trade_groups:
            type_label = "УралТрейд"
            status_parts.append(f"Штук: план {quantity_int}, факт {trade_codes_count}")
            receipt_firm = "УРАЛТРЕЙД"
            if check_boxes and cargo_places_int > 0:
                status_parts.append(f"Коробок: план {cargo_places_int}, факт {actual_packages_count}")
        else:
            type_label = "УралИмпорт"
            status_parts.append(f"Штук: план {quantity_int}, факт {import_codes_count}")
            receipt_firm = "УРАЛИМПОРТ"

        diff_str = " | ".join(status_parts)

        # Запись в файл-рассписку
        ws_receipt.append([article, quantity_int, total_codes_count, receipt_firm])

        if check_boxes and cargo_places_int > 0 and trade_groups:
            is_mismatch = (total_codes_count != quantity_int) or (actual_packages_count != cargo_places_int)
        else:
            is_mismatch = (total_codes_count != quantity_int)

        if is_mismatch:
            log_callback(f"[{type_label}] Артикул {article}: РАСХОЖДЕНИЕ! ({diff_str}). Но данные добавлены.")
            article_routing_summary.append((article, f"⚠️ Расхождение ({diff_str})"))
            mismatched_articles.append({
                "article": article, "expected": quantity_int, "actual": total_codes_count,
                "diff": abs(quantity_int - total_codes_count), "status": "Расхождение", "files": ", ".join(files_loaded_names)
            })
            bad_counter += 1
            stats_callback("bad", bad_counter)
        else:
            article_routing_summary.append((article, f"✅ {type_label} (Совпало)"))
            log_callback(f"[{type_label}] ... Артикул {article}: OK ({diff_str})")
            ok_counter += 1
            stats_callback("ok", ok_counter)

        if trade_groups:
            for file_name, p_groups in trade_groups.items():
                for pkg_code, codes_list in p_groups.items():
                    if not codes_list: continue
                    ws_trade[f"A{curr_trade}"] = "Код маркировки"
                    ws_trade[f"B{curr_trade}"] = file_name # Теперь без .xlsx
                    ws_trade[f"C{curr_trade}"] = name
                    ws_trade[f"D{curr_trade}"] = description
                    ws_trade[f"E{curr_trade}"] = str(cargo_places_int)
                    ws_trade[f"F{curr_trade}"] = diff_str
                    ws_trade[f"G{curr_trade}"] = str(quantity_int)
                    ws_trade[f"H{curr_trade}"] = unit
                    
                    display_pkg_code = remove_specific_brackets(pkg_code) if remove_brackets else pkg_code
                    ws_trade[f"I{curr_trade}"] = f"Код упаковки: {display_pkg_code}"
                    
                    curr_trade += 1
                    for code in codes_list:
                        ws_trade[f"A{curr_trade}"] = code
                        curr_trade += 1
                    curr_trade += 1 
                    
        if import_groups:
            for file_name, codes_list in import_groups.items():
                if not codes_list: continue
                ws_import[f"A{curr_import}"] = "Код маркировки"
                ws_import[f"B{curr_import}"] = file_name # Теперь без .xlsx
                ws_import[f"C{curr_import}"] = name
                ws_import[f"D{curr_import}"] = description
                ws_import[f"E{curr_import}"] = str(cargo_places_int)
                ws_import[f"F{curr_import}"] = diff_str
                ws_import[f"G{curr_import}"] = str(quantity_int)
                ws_import[f"H{curr_import}"] = unit
                curr_import += 1
                for code in codes_list:
                    ws_import[f"A{curr_import}"] = code
                    curr_import += 1
                curr_import += 1 

        progress_callback(index + 1, total_articles)

    # Выравнивание ширины столбцов
    for work_sheet in [ws_import, ws_trade, ws_receipt]:
        for col in work_sheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            column = col[0].column_letter
            for cell in col: cell.alignment = Alignment(vertical="top")
            work_sheet.column_dimensions[column].width = min(max_length + 5, 60)

    out_import = os.path.join(out_dir, "УралИмпорт.xlsx")
    out_trade = os.path.join(out_dir, "УралТрейд.xlsx")
    out_receipt = os.path.join(out_dir, "рассписка.xlsx")
    
    files_created = []
    
    if curr_import > 1:
        wb_import.save(out_import)
        files_created.append(out_import)
    if curr_trade > 1:
        wb_trade.save(out_trade)
        files_created.append(out_trade)
        
    # Сохраняем рассписку всегда
    wb_receipt.save(out_receipt)
    files_created.append(out_receipt)

    log_callback("\n============================================================")
    log_callback("ФИНАЛЬНЫЙ ОТЧЕТ ПО РАСХОЖДЕНИЯМ:")
    log_callback("============================================================\n")

    if not files_created:
        log_callback("[!] ВНИМАНИЕ: Не создано ни одного файла.")
    else:
        for f_path in files_created: log_callback(f"[✓] Успешно сохранен файл: {os.path.basename(f_path)}")

    if duplicate_summary:
        log_callback(f"\n[!] ОБНАРУЖЕНЫ СОВПАДЕНИЯ (ДУБЛИКАТЫ) КОДОВ МАРКИРОВКИ:")
        log_callback("    *Все типы дубликатов были полностью исключены из итоговых Excel-файлов.*")
    
    log_callback("\n============================================================")
    log_callback("СВОДНАЯ ТАБЛИЦА РАСПРЕДЕЛЕНИЯ АРТИКУЛОВ:")
    log_callback("============================================================")
    log_callback(f"{'АРТИКУЛ'.ljust(30)} | {'НАПРАВЛЕНИЕ / СТАТУС'}")
    log_callback("-" * 60)
    for art, route in article_routing_summary:
        log_callback(f"{art.ljust(30)} | {route}")
    log_callback("============================================================\n")

    log_callback("============================================================")
    log_callback(f"📊 ОБЩАЯ СУММА УНИКАЛЬНЫХ КОДОВ (ТРЕЙД + ИМПОРТ): {len(global_seen_codes)}")
    log_callback("============================================================\n")

    log_callback("============================================================")
    log_callback("НЕИСПОЛЬЗОВАННЫЕ ФАЙЛЫ В ПАПКЕ (ОСТАЛИСЬ НЕВЕРИФИЦИРОВАННЫМИ):")
    log_callback("============================================================")
    if unused_files:
        for f_unused in sorted(unused_files):
            log_callback(f" ⚠️ {f_unused}")
    else:
        log_callback(" 🎉 Отлично! Все Excel-файлы из папки были успешно задействованы.")
    log_callback("============================================================\n")


# =================================================================
# РЕЖИМ 2: БЫСТРАЯ ОЧИСТКА (Форматтер / Клинер)
# =================================================================
def process_cleaner_mode(c_folder, out_dir, log_callback, progress_callback, stats_callback, remove_brackets=False):
    all_files = [f for f in os.listdir(c_folder) if f.endswith(('.xlsx', '.xls')) and not f.startswith('~')]
    trade_files = [f for f in all_files if 'трейд' in f.lower() and 'clean' not in f.lower() and 'декларации' not in f.lower()]
    import_files = [f for f in all_files if 'импорт' in f.lower() and 'clean' not in f.lower() and 'декларации' not in f.lower()]

    total_files = len(trade_files) + len(import_files)
    stats_callback("total", total_files)
    
    if total_files == 0:
        log_callback("⚠️ В папке не найдены файлы, содержащие слова 'трейд' или 'импорт' (или они уже очищены).")
        return

    global_seen_marks = set()
    global_seen_packs = set()
    total_duplicates_removed = 0
    processed_count = 0

    # Создаем папку для Риммы
    rimma_dir = os.path.join(out_dir, "Для декларации")
    os.makedirs(rimma_dir, exist_ok=True)

    # ------------------ ОБРАБОТКА ТРЕЙДА ------------------
    for f_name in trade_files:
        log_callback(f"⏳ Обработка ТРЕЙД: {f_name}")
        file_path = os.path.join(c_folder, f_name)
        df = pd.read_excel(file_path, header=None)

        mark_codes = []
        pack_codes = []
        local_dups = 0

        for i in range(len(df)):
            if df.shape[1] > 0:
                val_a = clean_raw_value(df.iloc[i, 0])
                if val_a and len(val_a) >= 10:
                    if remove_brackets:
                        val_a = remove_specific_brackets(val_a)
                    if val_a not in global_seen_marks:
                        global_seen_marks.add(val_a)
                        mark_codes.append(val_a)
                    else:
                        local_dups += 1
                        total_duplicates_removed += 1
                        
            if df.shape[1] > 8:
                val_i = clean_raw_value(df.iloc[i, 8])
                if val_i and len(val_i) >= 10:
                    if remove_brackets:
                        val_i = remove_specific_brackets(val_i)
                    if val_i not in global_seen_packs:
                        global_seen_packs.add(val_i)
                        pack_codes.append(val_i)
                    else:
                        local_dups += 1
                        total_duplicates_removed += 1

        wb = Workbook()
        ws_mark = wb.active
        ws_mark.title = "Коды маркировки"
        for idx, code in enumerate(mark_codes, start=1):
            ws_mark[f"A{idx}"] = code
        ws_mark.column_dimensions['A'].width = 45

        if pack_codes:
            ws_pack = wb.create_sheet(title="Коды упаковки")
            for idx, code in enumerate(pack_codes, start=1):
                ws_pack[f"A{idx}"] = code
            ws_pack.column_dimensions['A'].width = 30

        out_name = f"CLEAN_{f_name}"
        wb.save(os.path.join(out_dir, out_name))
        log_callback(f"  ✅ Создан: {out_name} (Маркировок: {len(mark_codes)}, Упаковок: {len(pack_codes)} | Удалено дублей: {local_dups})")

        # --- ФАЙЛ ДЛЯ РИММЫ (УРАЛ ТРЕЙД) ---
        wb_rimma = Workbook()
        ws_rimma = wb_rimma.active
        ws_rimma.title = "Для декларации"
        
        for r_idx in range(len(df)):
            row_data = []
            for c_idx in range(df.shape[1]):
                cell_val = df.iloc[r_idx, c_idx]
                if pd.isna(cell_val):
                    row_data.append("")
                else:
                    # Применяем умную конвертацию в строку (чтобы не было 1200.0)
                    str_val = safe_str(cell_val)
                    str_val = re.sub(r'(?i)код упаковки\s*:\s*', '', str_val).strip()
                    row_data.append(str_val)
            ws_rimma.append(row_data)

        for col in ws_rimma.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws_rimma.column_dimensions[col[0].column_letter].width = min(max_length + 5, 50)

        base_name = os.path.splitext(f_name)[0]
        out_rimma = f"{base_name}_для_декларации.xlsx"
        # Сохраняем в отдельную папку!
        wb_rimma.save(os.path.join(rimma_dir, out_rimma))
        log_callback(f"  📁 Создан: {out_rimma} (В папке 'Для декларации')\n")
        
        processed_count += 1
        progress_callback(processed_count, total_files)
        stats_callback("ok", processed_count)
        stats_callback("dups", total_duplicates_removed)

    # ------------------ ОБРАБОТКА ИМПОРТА ------------------
    for f_name in import_files:
        log_callback(f"⏳ Обработка ИМПОРТ: {f_name}")
        file_path = os.path.join(c_folder, f_name)
        df = pd.read_excel(file_path, header=None)

        mark_codes = []
        local_dups = 0
        
        if len(df) > 0:
            for col_idx in range(df.shape[1]):
                col_has_codes = False
                for i in range(len(df)):
                    val_a = clean_raw_value(df.iloc[i, col_idx])
                    if val_a and len(val_a) >= 10:
                        col_has_codes = True
                        if remove_brackets:
                            val_a = remove_specific_brackets(val_a)
                        if val_a not in global_seen_marks:
                            global_seen_marks.add(val_a)
                            mark_codes.append(val_a)
                        else:
                            local_dups += 1
                            total_duplicates_removed += 1
                if col_has_codes:
                    break

        wb = Workbook()
        ws_mark = wb.active
        ws_mark.title = "Коды маркировки"
        for idx, code in enumerate(mark_codes, start=1):
            ws_mark[f"A{idx}"] = code
        ws_mark.column_dimensions['A'].width = 45

        out_name = f"CLEAN_{f_name}"
        wb.save(os.path.join(out_dir, out_name))
        log_callback(f"  ✅ Создан: {out_name} (Маркировок: {len(mark_codes)} | Удалено дублей: {local_dups})")

        # --- ФАЙЛ ДЛЯ РИММЫ (УРАЛ ИМПОРТ) ---
        wb_rimma = Workbook()
        ws_rimma = wb_rimma.active
        ws_rimma.title = "Для декларации"
        
        ws_rimma.append(["Код маркировки", "артикул", "наименование", "артикул 2", "места", "кол", "шт"])
        
        current_art = ""
        current_name = ""
        current_desc = ""
        current_places = ""
        current_qty = ""
        current_unit = ""
        is_first_code = False
        
        for r_idx in range(len(df)):
            # Применяем умную конвертацию в строку (отсекаем .0 у чисел)
            row_vals = [safe_str(x) for x in df.iloc[r_idx, :]]
            
            while row_vals and not row_vals[-1]:
                row_vals.pop()
                
            if not row_vals:
                continue
                
            val1 = row_vals[0].lower()
            
            if "код маркировки" in val1:
                current_art = row_vals[1] if len(row_vals) > 1 else ""
                current_name = row_vals[2] if len(row_vals) > 2 else ""
                current_desc = row_vals[3] if len(row_vals) > 3 else ""
                current_places = row_vals[4] if len(row_vals) > 4 else ""
                current_qty = row_vals[6] if len(row_vals) > 6 else ""
                current_unit = row_vals[7] if len(row_vals) > 7 else ""
                is_first_code = True
                continue
                
            code_val = row_vals[0]
            if code_val:
                if is_first_code:
                    ws_rimma.append([
                        code_val, current_art, current_name, current_desc,
                        current_places, current_qty, current_unit
                    ])
                    is_first_code = False
                else:
                    ws_rimma.append([
                        code_val, current_art, current_name, "", "", "", ""
                    ])

        for col in ws_rimma.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws_rimma.column_dimensions[col[0].column_letter].width = min(max_length + 5, 50)

        base_name = os.path.splitext(f_name)[0]
        out_rimma = f"{base_name}_для_декларации.xlsx"
        # Сохраняем в отдельную папку!
        wb_rimma.save(os.path.join(rimma_dir, out_rimma))
        log_callback(f"  📁 Создан: {out_rimma} (В папке 'Для декларации')\n")

        processed_count += 1
        progress_callback(processed_count, total_files)
        stats_callback("ok", processed_count)
        stats_callback("dups", total_duplicates_removed)

    log_callback(f"\n📊 ОБЩАЯ СУММА ОЧИЩЕННЫХ УНИКАЛЬНЫХ КОДОВ ЗА СЕССИЮ: {len(global_seen_marks)}")
    log_callback(f"🎉 ВСЕ ФАЙЛЫ ОЧИЩЕНЫ! (Всего отсеяно дубликатов за сессию: {total_duplicates_removed})")